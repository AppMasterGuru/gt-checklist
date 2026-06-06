"""
Tests for the 5 new features:
  warnings.py         — smart quote warnings
  provider_emails.py  — provider email generation
  sintad_export.py    — SINTAD Excel export
  wca_campaign.py     — WCA pilot campaign ZIP
  audit CSV export    — audit log CSV via Flask route

Target: 10 new tests bringing total to 74.
"""

from __future__ import annotations

import csv
import io
import json
import os
import tempfile
import zipfile

import pytest

# ── DB isolation (must patch before importing app modules) ────────────────────
_tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp_db.close()
os.environ["DB_PATH"] = _tmp_db.name

from core.db import get_connection, init_db  # noqa: E402
from core.warnings import check_quote_warnings, has_red_warnings  # noqa: E402
from core.provider_emails import generate_provider_emails  # noqa: E402
from core.sintad_export import generate_sintad_excel  # noqa: E402
from core.wca_campaign import generate_wca_campaign, get_pilot_agents, TRACKING_COLUMNS  # noqa: E402

import openpyxl  # noqa: E402


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def fresh_db():
    """Drop and recreate tables before every test for full isolation."""
    with get_connection() as conn:
        conn.executescript("""
            DROP TABLE IF EXISTS audit_log;
            DROP TABLE IF EXISTS quotes;
            DROP TABLE IF EXISTS ref_counters;
        """)
    init_db()
    yield


@pytest.fixture
def flask_client():
    from api.app import create_app
    app = create_app()
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


# ── Helper: minimal quote dict ────────────────────────────────────────────────

def _quote(overrides: dict | None = None) -> dict:
    q = {
        "reference_code": "26-05-001 Test FOB GT-PC",
        "client_name": "Test Client",
        "client_email": "test@example.com",
        "mode": "lcl",
        "incoterm": "FOB",
        "origin": "Lima, Perú",
        "destination": "Miami, USA",
        "cargo_description": "textiles generales",
        "weight_kg": 500.0,
        "volume_cbm": 2.0,
        "margin_pct": 0.20,
        "exchange_rate": 3.72,
        "staff_code": "GT-PC",
        "costeo_json": json.dumps({
            "flete_internacional_usd": 850.0,
            "visto_bueno_usd": 212.4,
            "customs_agent_usd": 100.0,
            "transport_usd": 53.76,
            "transport_soles": 200.0,
            "total_usd": 1216.16,
            "consolidator": "MSL",
            "customs_agent": "ALEFERO",
        }),
        "venta_json": json.dumps({
            "line_items": [
                {"description": "International Freight", "quantity": 1, "unit_price": 850.0, "total": 850.0},
                {"description": "Handling", "quantity": 1, "unit_price": 312.4, "total": 312.4},
                {"description": "Local Transport", "quantity": 1, "unit_price": 53.76, "total": 53.76},
            ],
            "total_usd": 1459.39,
            "margin_pct": 0.20,
            "validity_days": 15,
        }),
        "dimensions_json": json.dumps({"l": 100, "w": 80, "h": 60, "qty": 5}),
        "status": "PENDING",
    }
    if overrides:
        q.update(overrides)
    return q


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Warnings — RED block on low margin
# ═══════════════════════════════════════════════════════════════════════════════

def test_warnings_red_block_low_margin():
    """Margin below floor fires a yellow warning (soft — JP can override) and does NOT block approval."""
    q = _quote({"margin_pct": 0.08})
    warnings = check_quote_warnings(q)
    yellow_codes = [w["code"] for w in warnings if w["level"] == "yellow"]
    assert "MARGIN_BELOW_FLOOR" in yellow_codes
    # Must NOT block approval (no red for margin)
    assert has_red_warnings(warnings) is False


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Warnings — YELLOW on margin between 10% and 15%
# ═══════════════════════════════════════════════════════════════════════════════

def test_warnings_yellow_margin():
    """Margin 22% (between floor 20% and floor+5% 25%) fires MARGIN_LOW yellow, not red."""
    q = _quote({"margin_pct": 0.22})
    warnings = check_quote_warnings(q)
    codes     = [w["code"] for w in warnings]
    assert "MARGIN_LOW" in codes
    # has_red_warnings should be False for margin-only issue in low band
    margin_warns = [w for w in warnings if w["code"] in ("MARGIN_BELOW_FLOOR", "MARGIN_LOW")]
    assert any(w["level"] == "yellow" for w in margin_warns)
    assert has_red_warnings(warnings) is False


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Warnings — RED on dangerous goods without UN code
# ═══════════════════════════════════════════════════════════════════════════════

def test_warnings_dangerous_no_un_code():
    """Cargo described as 'peligrosa' without a UN code fires a red warning."""
    q = _quote({"cargo_description": "mercancía peligrosa — pintura industrial"})
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "DANGEROUS_NO_UN_CODE" in codes
    danger_warn = next(w for w in warnings if w["code"] == "DANGEROUS_NO_UN_CODE")
    assert danger_warn["level"] == "red"


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Warnings — RED for Farmex without OEA+BASC customs agent
# ═══════════════════════════════════════════════════════════════════════════════

def test_warnings_farmex_no_basc():
    """Farmex client without OEA+BASC agent must fire a red FARMEX_NO_BASC warning."""
    costeo = {
        "flete_internacional_usd": 800.0,
        "visto_bueno_usd": 0,
        "customs_agent_usd": 100.0,
        "transport_usd": 50.0,
        "transport_soles": 186.0,
        "total_usd": 950.0,
        "consolidator": "MSL",
        "customs_agent": "ALEFERO",   # NOT oea/basc
    }
    q = _quote({
        "client_name": "Farmex SA",
        "costeo_json": json.dumps(costeo),
    })
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "FARMEX_NO_BASC" in codes
    fw = next(w for w in warnings if w["code"] == "FARMEX_NO_BASC")
    assert fw["level"] == "red"


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Audit CSV export returns correct headers
# ═══════════════════════════════════════════════════════════════════════════════

def test_audit_export_csv(flask_client):
    """GET /audit/export.csv returns CSV with correct column headers."""
    resp = flask_client.get("/audit/export.csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.content_type
    lines = resp.data.decode("utf-8").splitlines()
    reader = csv.reader(lines)
    headers = next(reader)
    assert headers == ["id", "ts", "event_type", "quote_reference", "actor", "detail_json"]


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Provider emails — LCL generates 4 drafts
# ═══════════════════════════════════════════════════════════════════════════════

def test_provider_emails_lcl_generates_emails():
    """LCL mode produces one email draft per LCL provider; drafts are in Spanish."""
    q = _quote({"mode": "lcl"})
    emails = generate_provider_emails(q)
    assert len(emails) >= 1
    providers = {e["provider"] for e in emails}
    # Updated provider list: MSL, CRAFT, SACO, VANGUARD, ECU WORLDWIDE
    assert "MSL" in providers
    # All LCL drafts must be in Spanish and have to_emails key
    assert all(e["language"] == "es" for e in emails)
    assert all("to_emails" in e for e in emails)


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Provider emails — Aéreo generates 3 drafts in English
# ═══════════════════════════════════════════════════════════════════════════════

def test_provider_emails_aereo_generates_3():
    """Aéreo mode must produce 3 drafts (LAN, AA, United) in English."""
    q = _quote({"mode": "aereo"})
    emails = generate_provider_emails(q)
    assert len(emails) == 3
    providers = {e["provider"] for e in emails}
    assert providers == {"LAN Airlines", "American Airlines", "United Airlines"}
    assert all(e["language"] == "en" for e in emails)


# ═══════════════════════════════════════════════════════════════════════════════
# 8. SINTAD export produces 4 sheets
# ═══════════════════════════════════════════════════════════════════════════════

def test_sintad_export_has_4_sheets():
    """SINTAD Excel export must contain exactly 4 sheets."""
    q = _quote({"status": "APPROVED"})
    xlsx_bytes = generate_sintad_excel(q)
    assert len(xlsx_bytes) > 100  # non-empty
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert len(wb.sheetnames) == 4
    assert wb.sheetnames == ["Datos Generales", "Costos", "Venta", "Staff"]


# ═══════════════════════════════════════════════════════════════════════════════
# 9. WCA campaign ZIP contains correct number of email files
# ═══════════════════════════════════════════════════════════════════════════════

def test_wca_campaign_generates_correct_count(monkeypatch):
    """WCA campaign ZIP must contain N email files plus deck and CSV."""
    monkeypatch.setenv("WCA_SENDER_EMAIL", "comercial@gt.com.pe")
    n = 12
    zip_bytes = generate_wca_campaign("Colombia", "textiles", n, "es", "Renato Alvarez")
    assert len(zip_bytes) > 100
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
    email_files = [f for f in names if f.startswith("emails/")]
    assert len(email_files) == n
    assert "capability_deck.txt" in names
    # Tracking CSV present
    csv_files = [f for f in names if f.endswith(".csv")]
    assert len(csv_files) == 1


# ═══════════════════════════════════════════════════════════════════════════════
# 10. WCA tracking CSV has correct columns
# ═══════════════════════════════════════════════════════════════════════════════

def test_wca_tracking_csv_columns(monkeypatch):
    """Tracking CSV in WCA ZIP must have exactly the specified column headers."""
    monkeypatch.setenv("WCA_SENDER_EMAIL", "comercial@gt.com.pe")
    zip_bytes = generate_wca_campaign("Alemania", "farmacéutico", 10, "de", "Renato Alvarez")
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        csv_name = next(f for f in zf.namelist() if f.endswith(".csv"))
        raw = zf.read(csv_name).decode("utf-8")
    reader = csv.reader(io.StringIO(raw))
    headers = next(reader)
    assert headers == TRACKING_COLUMNS
    # Should have 10 data rows
    rows = list(reader)
    assert len(rows) == 10


# ═══════════════════════════════════════════════════════════════════════════════
# 11. send_provider_email() — stub mode returns (True, msg) and logs audit event
# ═══════════════════════════════════════════════════════════════════════════════

def test_send_provider_email_stub_returns_true():
    """send_provider_email in stub mode (no Graph creds) must return (True, ...)."""
    from core.email_sender import send_provider_email
    from core.db import get_audit_trail, init_db

    init_db()
    ok, msg = send_provider_email(
        ref_code="26-05-TEST",
        provider="MSL",
        to="msl@example.com",
        subject="Test rate request",
        body="Please quote Lima–Hamburg LCL.",
        actor="abel",
    )
    assert ok is True
    assert "MSL" in msg or "msl@example.com" in msg

    # Audit event logged
    trail = get_audit_trail("26-05-TEST")
    events = [e["event_type"] for e in trail]
    assert "PROVIDER_EMAIL_SENT" in events


# ═══════════════════════════════════════════════════════════════════════════════
# 12. POST /quote/<ref>/provider-emails/send — redirects (302) back to review page
# ═══════════════════════════════════════════════════════════════════════════════

def test_provider_emails_send_route_redirects(flask_client):
    """POST provider-emails/send must return 302 for a valid quote ref."""
    from core.db import get_connection, init_db

    init_db()
    # Insert a minimal quote so the route can find it
    with get_connection() as conn:
        conn.execute("""
            INSERT OR IGNORE INTO quotes
              (reference_code, client_name, client_email,
               mode, incoterm, origin, destination,
               cargo_description, weight_kg, volume_cbm,
               margin_pct, exchange_rate, staff_code, status)
            VALUES
              ('26-05-PROV-TEST', 'Test Client', 'test@example.com',
               'lcl', 'FOB', 'Lima', 'Miami',
               'general cargo', 500.0, 2.0,
               0.20, 3.72, 'GT-PC', 'PENDING')
        """)

    resp = flask_client.post(
        "/quote/26-05-PROV-TEST/provider-emails/send",
        follow_redirects=False,
    )
    assert resp.status_code == 302


# ═══════════════════════════════════════════════════════════════════════════════
# Fix 6 — Configurable margin floor (MIN_MARGIN_PCT env var)
# ═══════════════════════════════════════════════════════════════════════════════

def test_fix6_below_floor_yellow_warning():
    """Margin below MIN_MARGIN_PCT (20%) fires MARGIN_BELOW_FLOOR as yellow."""
    q = _quote({"margin_pct": 0.10})
    warnings = check_quote_warnings(q)
    codes  = [w["code"]  for w in warnings]
    levels = {w["level"] for w in warnings}
    assert "MARGIN_BELOW_FLOOR" in codes
    below = next(w for w in warnings if w["code"] == "MARGIN_BELOW_FLOOR")
    assert below["level"] == "yellow"
    # Must NOT set has_red — approve button stays enabled
    assert has_red_warnings(warnings) is False


def test_fix6_above_floor_no_margin_warning():
    """Margin well above floor produces no margin-related warning."""
    q = _quote({"margin_pct": 0.30})
    warnings = check_quote_warnings(q)
    margin_codes = [w["code"] for w in warnings
                    if w["code"] in ("MARGIN_BELOW_FLOOR", "MARGIN_LOW")]
    assert margin_codes == []


def test_fix6_above_band_no_warning():
    """Margin at 25%+ (floor+5%) produces no margin warning at all."""
    q = _quote({"margin_pct": 0.25})
    warnings = check_quote_warnings(q)
    margin_codes = [w["code"] for w in warnings
                    if w["code"] in ("MARGIN_BELOW_FLOOR", "MARGIN_LOW")]
    assert margin_codes == []


def test_fix6_approval_not_blocked_below_floor():
    """has_red_warnings returns False even when margin is below floor (non-blocking)."""
    q = _quote({"margin_pct": 0.05})
    warnings = check_quote_warnings(q)
    # Margin warning is yellow — approval gate must NOT be disabled
    assert has_red_warnings(warnings) is False


# ═══════════════════════════════════════════════════════════════════════════════
# WCA personalised campaign — 4 new tests (Italy + Textiles pilot)
# ═══════════════════════════════════════════════════════════════════════════════

def test_wca_personalised_no_placeholders(monkeypatch):
    """Every email in the Italy+Textiles ZIP must have zero unfilled placeholders."""
    monkeypatch.setenv("WCA_SENDER_EMAIL", "comercial@gt.com.pe")
    agents = get_pilot_agents("Italia", "textiles")
    assert len(agents) == 25, "Italy+Textiles pilot list must have exactly 25 agents"

    zip_bytes = generate_wca_campaign(
        "Italia", "textiles", language="en",
        sender_name="Renato Alvarez", agents=agents,
    )
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        email_files = sorted(f for f in zf.namelist() if f.startswith("emails/"))
        assert len(email_files) == 25
        # Spot-check files 1, 13, and 25 (first, middle, last)
        for fname in [email_files[0], email_files[12], email_files[24]]:
            content = zf.read(fname).decode("utf-8")
            assert "[AGENT_NAME]"    not in content, f"{fname} still contains [AGENT_NAME]"
            assert "[AGENT_COMPANY]" not in content, f"{fname} still contains [AGENT_COMPANY]"


def test_wca_personalised_file_naming(monkeypatch):
    """Email files must be named email_NNN_company_slug.txt, not email_NNN.txt."""
    monkeypatch.setenv("WCA_SENDER_EMAIL", "comercial@gt.com.pe")
    agents = get_pilot_agents("Italia", "textiles")

    zip_bytes = generate_wca_campaign(
        "Italia", "textiles", language="en",
        sender_name="Renato Alvarez", agents=agents,
    )
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        email_files = [f for f in zf.namelist() if f.startswith("emails/")]

    # Every file must have a company slug after the sequence number
    for fname in email_files:
        basename = fname.split("/")[-1]         # e.g. email_001_pellegrini_forwarding_s_r_l_.txt
        parts = basename.split("_")
        assert len(parts) >= 3, f"Expected email_NNN_slug.txt format, got: {fname}"
        assert parts[0] == "email"
        assert parts[1].isdigit()

    # First file should reference the first agent's company
    first = sorted(email_files)[0]
    assert "pellegrini" in first, f"First file should contain 'pellegrini', got: {first}"


def test_wca_personalised_csv_prefilled(monkeypatch):
    """Tracking CSV must have one row per agent with name, company, and email pre-filled."""
    monkeypatch.setenv("WCA_SENDER_EMAIL", "comercial@gt.com.pe")
    agents = get_pilot_agents("Italia", "textiles")

    zip_bytes = generate_wca_campaign(
        "Italia", "textiles", language="en",
        sender_name="Renato Alvarez", agents=agents,
    )
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        csv_name = next(f for f in zf.namelist() if f.endswith(".csv"))
        raw = zf.read(csv_name).decode("utf-8")

    reader = csv.reader(io.StringIO(raw))
    headers = next(reader)
    assert headers == TRACKING_COLUMNS

    rows = list(reader)
    assert len(rows) == 25

    # Spot-check: first row must have real agent data, not placeholders
    first = rows[0]
    assert first[0] == "Marco Pellegrini"
    assert first[1] == "Pellegrini Forwarding S.r.l."
    assert first[2] == "m.pellegrini@pellegrini-fwd.it"
    assert first[3] == ""   # Sent — blank
    assert first[4] == ""   # Reply — blank
    assert first[5] == ""   # Notes — blank

    # No row should contain unfilled bracket placeholders
    for row in rows:
        for cell in row:
            assert "[Agente" not in cell,  f"Unfilled placeholder in CSV row: {row}"
            assert "[Empresa]" not in cell, f"Unfilled placeholder in CSV row: {row}"


def test_wca_sender_email_required(monkeypatch):
    """generate_wca_campaign must raise ValueError with actionable message when WCA_SENDER_EMAIL is unset."""
    monkeypatch.delenv("WCA_SENDER_EMAIL", raising=False)

    with pytest.raises(ValueError) as exc_info:
        generate_wca_campaign("Italia", "textiles", 10, "en", "Renato Alvarez")

    assert "WCA_SENDER_EMAIL" in str(exc_info.value)


# ── /run-listener route ───────────────────────────────────────────────────────

def test_run_listener_disabled_returns_correct_json(flask_client, monkeypatch):
    """POST /run-listener with LISTENER_ENABLED unset returns LISTENER_DISABLED status."""
    monkeypatch.delenv("LISTENER_ENABLED", raising=False)
    resp = flask_client.post("/run-listener")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "LISTENER_DISABLED"
    assert data["emails_processed"] == 0
    assert data["acks_queued"] == 0
    assert "ts" in data


def test_run_listener_enabled_returns_ok_json(flask_client, monkeypatch):
    """POST /run-listener with LISTENER_ENABLED=true returns OK and correct counts."""
    monkeypatch.setenv("LISTENER_ENABLED", "true")
    fake_results = [
        {"ack_queued": True},
        {"ack_queued": False},
        {"ack_queued": True},
    ]
    monkeypatch.setattr("api.routes.process_inbound_emails", lambda auto_ack: fake_results)
    resp = flask_client.post("/run-listener")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "OK"
    assert data["emails_processed"] == 3
    assert data["acks_queued"] == 2
    assert "ts" in data


def test_run_listener_disabled_creates_audit_entry(flask_client, monkeypatch):
    """POST /run-listener when disabled writes a LISTENER_POLL audit row with LISTENER_DISABLED."""
    monkeypatch.delenv("LISTENER_ENABLED", raising=False)
    flask_client.post("/run-listener")
    with get_connection() as conn:
        row = conn.execute(
            "SELECT event_type, detail_json FROM audit_log WHERE event_type='LISTENER_POLL' ORDER BY id DESC LIMIT 1"
        ).fetchone()
    assert row is not None
    detail = json.loads(row["detail_json"])
    assert detail["status"] == "LISTENER_DISABLED"


def test_run_listener_enabled_creates_audit_entry(flask_client, monkeypatch):
    """POST /run-listener when enabled writes a LISTENER_POLL audit row with OK status."""
    monkeypatch.setenv("LISTENER_ENABLED", "true")
    monkeypatch.setattr("api.routes.process_inbound_emails", lambda auto_ack: [{"ack_queued": True}])
    flask_client.post("/run-listener")
    with get_connection() as conn:
        row = conn.execute(
            "SELECT event_type, detail_json FROM audit_log WHERE event_type='LISTENER_POLL' ORDER BY id DESC LIMIT 1"
        ).fetchone()
    assert row is not None
    detail = json.loads(row["detail_json"])
    assert detail["status"] == "OK"
    assert detail["emails_processed"] == 1
    assert detail["acks_queued"] == 1


# ── Additional warnings coverage ─────────────────────────────────────────────

def test_warnings_perishable_no_temp():
    """Perishable cargo without temperature info fires PERISHABLE_NO_TEMP yellow."""
    q = _quote({"cargo_description": "flores frescas de exportación"})
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "PERISHABLE_NO_TEMP" in codes
    warn = next(w for w in warnings if w["code"] == "PERISHABLE_NO_TEMP")
    assert warn["level"] == "yellow"


def test_warnings_perishable_with_temp_no_warning():
    """Perishable cargo with temperature info does NOT fire PERISHABLE_NO_TEMP."""
    q = _quote({"cargo_description": "flores frescas temperatura controlada 4°C"})
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "PERISHABLE_NO_TEMP" not in codes


def test_warnings_no_client_email():
    """Quote missing client_email fires NO_CLIENT_EMAIL yellow warning."""
    q = _quote({"client_email": ""})
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "NO_CLIENT_EMAIL" in codes
    warn = next(w for w in warnings if w["code"] == "NO_CLIENT_EMAIL")
    assert warn["level"] == "yellow"


def test_warnings_high_density():
    """Weight/volume giving density > 2000 kg/m³ fires HIGH_DENSITY yellow."""
    q = _quote({"weight_kg": 10000.0, "volume_cbm": 1.0})
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "HIGH_DENSITY" in codes


def test_warnings_low_density():
    """Weight/volume giving density < 50 kg/m³ fires LOW_DENSITY yellow."""
    q = _quote({"weight_kg": 10.0, "volume_cbm": 5.0})
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "LOW_DENSITY" in codes


def test_warnings_zero_venta():
    """venta_json total_usd = 0 fires ZERO_VENTA red warning."""
    q = _quote({})
    q["venta_json"] = {"total_usd": 0, "line_items": [], "margin_pct": 0.25, "validity_days": 15}
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "ZERO_VENTA" in codes
    warn = next(w for w in warnings if w["code"] == "ZERO_VENTA")
    assert warn["level"] == "red"


def test_warnings_zero_flete():
    """costeo_json flete_internacional_usd = 0 fires ZERO_FLETE yellow warning."""
    q = _quote({})
    q["costeo_json"] = {
        "flete_internacional_usd": 0, "visto_bueno_usd": 50,
        "customs_agent_usd": 100, "transport_usd": 80,
        "transport_soles": 300, "total_usd": 230, "customs_agent": "OEA",
    }
    warnings = check_quote_warnings(q)
    codes = [w["code"] for w in warnings]
    assert "ZERO_FLETE" in codes
    warn = next(w for w in warnings if w["code"] == "ZERO_FLETE")
    assert warn["level"] == "yellow"


def test_has_red_warnings_true():
    """has_red_warnings returns True when any warning has level red."""
    ws = [{"level": "yellow", "code": "A", "message": "x"},
          {"level": "red",    "code": "B", "message": "y"}]
    assert has_red_warnings(ws) is True


def test_has_red_warnings_false_all_yellow():
    """has_red_warnings returns False when all warnings are yellow."""
    ws = [{"level": "yellow", "code": "A", "message": "x"},
          {"level": "yellow", "code": "C", "message": "z"}]
    assert has_red_warnings(ws) is False


# ── email_sender coverage ─────────────────────────────────────────────────────

def test_send_quote_email_stub_returns_true():
    """send_quote_email in stub mode returns (True, message)."""
    from core.email_sender import send_quote_email
    from core.db import init_db
    init_db()
    ok, msg = send_quote_email(
        ref_code="26-05-QE-TEST",
        quote_id=999,
        customer_email="cliente@example.com",
        customer_name="Test SA",
        actor="abel",
    )
    assert ok is True
    assert "26-05-QE-TEST" in msg


def test_send_quote_email_logs_audit():
    """send_quote_email logs QUOTE_SENT to audit trail."""
    from core.email_sender import send_quote_email
    from core.db import init_db, get_audit_trail
    init_db()
    send_quote_email(
        ref_code="26-05-QE-AUDIT",
        quote_id=998,
        customer_email="x@example.com",
        customer_name="Audit SA",
        actor="jp",
    )
    trail = get_audit_trail("26-05-QE-AUDIT")
    assert any(e["event_type"] == "QUOTE_SENT" for e in trail)


def test_send_acknowledgment_email_stub_returns_true():
    """send_acknowledgment_email in stub mode returns (True, message)."""
    from core.email_sender import send_acknowledgment_email
    from core.db import init_db
    init_db()
    ok, msg = send_acknowledgment_email(
        recipient_email="ack@example.com",
        recipient_name="Test User",
        subject="Acuse de Recibo",
        ack_text="Hemos recibido su solicitud.",
        actor="system",
    )
    assert ok is True
    assert "ack@example.com" in msg


def test_credentials_rotated_is_bool():
    """CREDENTIALS_ROTATED must be a bool — routes import it for template context."""
    from core.email_sender import CREDENTIALS_ROTATED
    assert isinstance(CREDENTIALS_ROTATED, bool)
