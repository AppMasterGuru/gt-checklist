"""
Tests for core/sintad_export.py — SINTAD Excel pre-fill generator.

generate_sintad_excel(quote) → bytes (openpyxl workbook)

4 sheets: Datos Generales, Costos, Venta, Staff
Staff routing: Jean Paul for GT-PC/GT-WCA, Renato for others
Operativo:    Junior Loa (export), Robin Lujan (import)
Export detection: Lima/Peru/Callao in origin → export
"""

from __future__ import annotations

import io
import json

import openpyxl

from core.sintad_export import generate_sintad_excel, _is_export, _staff


# ── Fixtures ─────────────────────────────────────────────────────────────────

def _quote(**overrides) -> dict:
    base = {
        "reference_code": "26-05-TEST",
        "client_name": "Test SA",
        "origin": "Lima, Peru",
        "destination": "Hamburg, Germany",
        "incoterm": "FOB",
        "mode": "lcl",
        "cargo_description": "harina de quinua",
        "weight_kg": 3200.0,
        "volume_cbm": 4.8,
        "exchange_rate": 3.72,
        "staff_code": "GT-PC",
        "costeo_json": json.dumps({
            "flete_internacional_usd": 850.0,
            "visto_bueno_usd": 45.0,
            "customs_agent_usd": 120.0,
            "transport_usd": 95.0,
            "transport_soles": 354.0,
            "total_usd": 1110.0,
            "consolidator": "MSL",
            "customs_agent": "Ausa",
        }),
        "venta_json": json.dumps({
            "total_usd": 1388.0,
            "margin_pct": 0.20,
            "validity_days": 15,
            "line_items": [
                {"description": "Flete LCL", "quantity": 1, "unit_price": 1388.0, "total": 1388.0}
            ],
        }),
        "dimensions_json": None,
    }
    base.update(overrides)
    return base


def _load_wb(quote_dict) -> openpyxl.Workbook:
    raw = generate_sintad_excel(quote_dict)
    return openpyxl.load_workbook(io.BytesIO(raw))


# ── Basic output ──────────────────────────────────────────────────────────────

class TestSintadBasic:
    def test_returns_bytes(self):
        result = generate_sintad_excel(_quote())
        assert isinstance(result, bytes) and len(result) > 0

    def test_has_exactly_4_sheets(self):
        wb = _load_wb(_quote())
        assert len(wb.sheetnames) == 4

    def test_sheet_names(self):
        wb = _load_wb(_quote())
        assert wb.sheetnames == ["Datos Generales", "Costos", "Venta", "Staff"]

    def test_reference_appears_in_datos_generales(self):
        wb = _load_wb(_quote(reference_code="26-05-REF-123"))
        ws = wb["Datos Generales"]
        found = any(
            "26-05-REF-123" in str(cell.value or "")
            for row in ws.iter_rows() for cell in row
        )
        assert found

    def test_empty_quote_does_not_crash(self):
        result = generate_sintad_excel({})
        assert isinstance(result, bytes)


# ── Export / import detection ─────────────────────────────────────────────────

class TestIsExport:
    def test_lima_origin_is_export(self):
        assert _is_export({"origin": "Lima, Peru"}) is True

    def test_callao_origin_is_export(self):
        assert _is_export({"origin": "Callao"}) is True

    def test_hamburg_origin_is_import(self):
        assert _is_export({"origin": "Hamburg, Germany"}) is False

    def test_empty_origin_is_import(self):
        assert _is_export({"origin": ""}) is False


# ── Staff routing ─────────────────────────────────────────────────────────────

class TestStaffRouting:
    def test_gt_pc_gives_jean_paul(self):
        staff = _staff({"staff_code": "GT-PC", "origin": "Hamburg"})
        assert staff["Ejecutivo Comercial"] == "Jean Paul"

    def test_gt_wca_gives_jean_paul(self):
        staff = _staff({"staff_code": "GT-WCA", "origin": "Hamburg"})
        assert staff["Ejecutivo Comercial"] == "Jean Paul"

    def test_other_code_gives_renato(self):
        staff = _staff({"staff_code": "GT-OTHER", "origin": "Hamburg"})
        assert staff["Ejecutivo Comercial"] == "Renato"

    def test_export_gives_junior_loa(self):
        staff = _staff({"staff_code": "GT-PC", "origin": "Lima"})
        assert staff["Operativo"] == "Junior Loa"

    def test_import_gives_robin_lujan(self):
        staff = _staff({"staff_code": "GT-PC", "origin": "Hamburg"})
        assert staff["Operativo"] == "Robin Lujan"

    def test_customer_service_always_paulo(self):
        staff = _staff({"staff_code": "GT-PC", "origin": "Lima"})
        assert staff["Customer Service"] == "Paulo Díaz"

    def test_supervisor_always_kristel(self):
        staff = _staff({"staff_code": "GT-OTHER", "origin": "Lima"})
        assert staff["Supervisor"] == "Kristel"
