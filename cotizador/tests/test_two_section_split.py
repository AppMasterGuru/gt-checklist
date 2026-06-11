"""
Tests for the two-section split: Costos de Flete Internacional / Gastos Locales.

Covers:
  - is_international / is_local / igv_applicable flags on venta line items
  - Section split in pdf_generator (flete table / local table HTML)
  - SINTAD Costos + Venta sheets: section header rows present
  - SINTAD IGV column on local items
  - Extra coloader items default to is_international
"""

from __future__ import annotations

import io
import json
import os
import tempfile

import openpyxl
import pytest

_tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp_db.close()
os.environ.setdefault("DB_PATH", _tmp_db.name)

from core.db import get_connection, init_db  # noqa: E402
from core.pdf_generator import _build_flete_table, _build_local_table, render_html  # noqa: E402
from core.sintad_export import generate_sintad_excel  # noqa: E402


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def fresh_db():
    with get_connection() as conn:
        conn.executescript("""
            DROP TABLE IF EXISTS audit_log;
            DROP TABLE IF EXISTS quotes;
            DROP TABLE IF EXISTS ref_counters;
            DROP TABLE IF EXISTS providers;
            DROP TABLE IF EXISTS provider_replies;
            DROP TABLE IF EXISTS credit_registry;
        """)
    init_db()
    yield


@pytest.fixture
def app():
    from api.app import create_app
    a = create_app()
    a.config["TESTING"] = True
    return a


@pytest.fixture
def client(app):
    with app.test_client() as c:
        yield c


_BASE_FORM = {
    "client_name": "SplitTest SA",
    "client_email": "test@split.com",
    "mode": "lcl",
    "incoterm": "FOB",
    "origin": "Lima, Peru",
    "destination": "Hamburg, Germany",
    "cargo_description": "harina de quinua",
    "weight": "500",
    "weight_unit": "kg",
    "volume_cbm": "2.0",
    "flete_lcl": "200.00",
    "consolidator": "MSL",
    "staff_code": "GT-PC",
    "language": "es",
    "requester_type": "cliente",
}


def _post_quote(client, overrides=None):
    from urllib.parse import unquote
    data = {**_BASE_FORM, **(overrides or {})}
    resp = client.post("/quote/new", data=data, follow_redirects=False)
    assert resp.status_code == 302
    ref = unquote(resp.headers["Location"].split("/quote/")[-1].rstrip("/"))
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM quotes WHERE reference_code=?", (ref,)).fetchone()
    return dict(row)


def _sintad_quote(**overrides) -> dict:
    base = {
        "reference_code": "26-05-TEST",
        "client_name": "Test SA",
        "origin": "Lima, Peru",
        "destination": "Hamburg, Germany",
        "incoterm": "FOB",
        "mode": "lcl",
        "cargo_description": "harina de quinua",
        "weight_kg": 3200.0,
        "volume_cbm": 4.8,
        "exchange_rate": 3.72,
        "staff_code": "GT-PC",
        "costeo_json": json.dumps({
            "flete_internacional_usd": 850.0,
            "visto_bueno_usd": 45.0,
            "customs_agent_usd": 120.0,
            "transport_usd": 95.0,
            "transport_soles": 354.0,
            "handling_aereo_usd": 0.0,
            "total_usd": 1110.0,
            "consolidator": "MSL",
            "customs_agent": "Ausa",
        }),
        "venta_json": json.dumps({
            "total_usd": 1388.0,
            "margin_pct": 0.20,
            "validity_days": 15,
            "line_items": [
                {"description": "International Freight", "quantity": 1,
                 "unit_price": 1020.0, "total": 1020.0,
                 "is_international": True, "is_local": False, "igv_applicable": False},
                {"description": "Visto Bueno", "quantity": 1,
                 "unit_price": 54.0, "total": 54.0,
                 "is_international": False, "is_local": True, "igv_applicable": True},
                {"description": "Agente de Aduana", "quantity": 1,
                 "unit_price": 144.0, "total": 144.0,
                 "is_international": False, "is_local": True, "igv_applicable": True},
                {"description": "Transporte Local", "quantity": 1,
                 "unit_price": 114.0, "total": 114.0,
                 "is_international": False, "is_local": True, "igv_applicable": True},
            ],
        }),
        "dimensions_json": None,
    }
    base.update(overrides)
    return base


def _load_wb(quote_dict) -> openpyxl.Workbook:
    raw = generate_sintad_excel(quote_dict)
    return openpyxl.load_workbook(io.BytesIO(raw))


# ── Section flags on venta line items ────────────────────────────────────────

class TestSectionFlags:
    def test_flete_item_is_international(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        flete = next(i for i in venta["line_items"] if i["description"] == "International Freight")
        assert flete["is_international"] is True
        assert flete["is_local"] is False
        assert flete["igv_applicable"] is False

    def test_local_items_have_local_flag(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        local_items = [i for i in venta["line_items"] if i.get("is_local")]
        assert len(local_items) >= 1
        for item in local_items:
            assert item["is_local"] is True
            assert item["is_international"] is False
            assert item["igv_applicable"] is True

    def test_agente_aduana_is_local(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        agente = next(i for i in venta["line_items"] if i["description"] == "Agente de Aduana")
        assert agente["is_local"] is True
        assert agente["igv_applicable"] is True

    def test_transporte_local_is_local(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        transport = next((i for i in venta["line_items"] if i["description"] == "Transporte Local"), None)
        if transport:
            assert transport["is_local"] is True

    def test_thc_is_international(self, client):
        q = _post_quote(client, {
            "thc_rate": "12", "thc_min": "30",
            "volume_cbm": "3.0", "weight": "1000", "weight_unit": "kg",
        })
        venta = json.loads(q["venta_json"])
        thc = next((i for i in venta["line_items"] if "THC" in i["description"]), None)
        if thc:
            assert thc["is_international"] is True
            assert thc["is_local"] is False

    def test_no_item_missing_flags(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        for item in venta["line_items"]:
            assert "is_local" in item
            assert "is_international" in item
            assert "igv_applicable" in item

    def test_extra_item_defaults_to_is_international(self, client):
        q = _post_quote(client, {
            "extra_items_json": '[{"concept":"Pick Up","valor":150,"factor":null,"min_usd":null}]',
        })
        venta = json.loads(q["venta_json"])
        pick_up = next(i for i in venta["line_items"] if i["description"] == "Pick Up")
        assert pick_up["is_international"] is True
        assert pick_up["is_local"] is False

    def test_extra_item_with_factor_defaults_to_is_international(self, client):
        q = _post_quote(client, {
            "volume_cbm": "3.0", "weight": "1000", "weight_unit": "kg",
            "extra_items_json": '[{"concept":"Surcharge","valor":10,"factor":3,"min_usd":null}]',
        })
        venta = json.loads(q["venta_json"])
        surcharge = next(i for i in venta["line_items"] if i["description"] == "Surcharge")
        assert surcharge["is_international"] is True

    def test_visto_bueno_present_when_lcl(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        local_descs = [i["description"] for i in venta["line_items"] if i.get("is_local")]
        assert "Visto Bueno" in local_descs

    def test_intl_items_have_no_igv(self, client):
        q = _post_quote(client)
        venta = json.loads(q["venta_json"])
        for item in venta["line_items"]:
            if item.get("is_international"):
                assert item["igv_applicable"] is False

    def test_dynamic_mode_also_has_local_items(self, client):
        q = _post_quote(client, {
            "extra_items_json": '[{"concept":"Pick Up","valor":100,"factor":null,"min_usd":null}]',
        })
        venta = json.loads(q["venta_json"])
        local_items = [i for i in venta["line_items"] if i.get("is_local")]
        assert len(local_items) >= 1


# ── pdf_generator section builders ───────────────────────────────────────────

_INTL = [
    {"description": "International Freight", "quantity": 1, "unit_price": 1020.0,
     "total": 1020.0, "is_international": True, "is_local": False, "igv_applicable": False},
]
_LOCAL = [
    {"description": "Visto Bueno", "quantity": 1, "unit_price": 54.0,
     "total": 54.0, "is_international": False, "is_local": True, "igv_applicable": True},
    {"description": "Transporte Local", "quantity": 1, "unit_price": 114.0,
     "total": 114.0, "is_international": False, "is_local": True, "igv_applicable": True},
]


class TestPdfSectionBuilders:
    def test_flete_table_section_header_es(self):
        html = _build_flete_table(_INTL, "es")
        assert "Costos de Flete Internacional" in html

    def test_flete_table_section_header_en(self):
        html = _build_flete_table(_INTL, "en")
        assert "International Freight Charges" in html

    def test_flete_table_has_subtotal(self):
        html = _build_flete_table(_INTL, "es")
        assert "Subtotal Flete" in html
        assert "1,020.00" in html

    def test_local_table_section_header_es(self):
        html = _build_local_table(_LOCAL, "es")
        assert "Gastos Locales" in html
        assert "IGV 18%" in html

    def test_local_table_section_header_en(self):
        html = _build_local_table(_LOCAL, "en")
        assert "Local Charges" in html
        assert "VAT 18%" in html

    def test_local_table_igv_amount(self):
        html = _build_local_table(_LOCAL, "es")
        neto = 54.0 + 114.0
        assert f"{neto * 0.18:,.2f}" in html

    def test_local_table_total_incl_igv(self):
        html = _build_local_table(_LOCAL, "es")
        total = (54.0 + 114.0) * 1.18
        assert f"{total:,.2f}" in html

    def test_empty_local_returns_empty(self):
        assert _build_local_table([], "es") == ""

    def test_empty_flete_returns_empty(self):
        assert _build_flete_table([], "es") == ""

    def test_render_html_contains_both_sections(self):
        venta = {"line_items": _INTL + _LOCAL, "total_usd": 1388.0, "margin_pct": 0.20}
        meta  = {"reference": "26-TEST", "client_name": "Test", "language": "es",
                 "staff_code": "GT-PC", "origin": "Lima", "destination": "Hamburg",
                 "incoterm": "FOB", "mode": "LCL", "exchange_rate": 3.72}
        html = render_html(venta, meta)
        assert "Costos de Flete Internacional" in html
        assert "Gastos Locales" in html

    def test_render_html_local_subtotal_includes_igv(self):
        venta = {"line_items": _INTL + _LOCAL, "total_usd": 0, "margin_pct": 0.20}
        meta  = {"reference": "26-TEST", "client_name": "Test", "language": "es",
                 "staff_code": "GT-PC", "origin": "Lima", "destination": "Hamburg",
                 "incoterm": "FOB", "mode": "LCL", "exchange_rate": 3.72}
        html = render_html(venta, meta)
        local_neto = 54.0 + 114.0
        assert f"{local_neto * 1.18:,.2f}" in html

    def test_render_html_en_uses_english_labels(self):
        venta = {"line_items": _INTL + _LOCAL, "total_usd": 0, "margin_pct": 0.20}
        meta  = {"reference": "26-TEST", "client_name": "Test", "language": "en",
                 "staff_code": "GT-PC", "origin": "Lima", "destination": "Hamburg",
                 "incoterm": "FOB", "mode": "LCL", "exchange_rate": 3.72}
        html = render_html(venta, meta)
        assert "International Freight Charges" in html
        assert "Local Charges" in html


# ── SINTAD section headers ────────────────────────────────────────────────────

class TestSintadSectionHeaders:
    def test_costos_has_flete_section_header(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Costos"]
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 30)]
        assert any("FLETE INTERNACIONAL" in v.upper() for v in values)

    def test_costos_has_locales_section_header(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Costos"]
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 30)]
        assert any("GASTOS LOCALES" in v.upper() for v in values)

    def test_venta_has_flete_section_header(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Venta"]
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 40)]
        assert any("FLETE INTERNACIONAL" in v.upper() for v in values)

    def test_venta_has_locales_section_header(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Venta"]
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 40)]
        assert any("GASTOS LOCALES" in v.upper() for v in values)


# ── SINTAD IGV column ─────────────────────────────────────────────────────────

class TestSintadIgvColumn:
    def test_costos_has_igv_column_header(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Costos"]
        all_vals = [
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, 30) for c in range(1, 6)
        ]
        assert any("IGV" in v for v in all_vals)

    def test_venta_has_igv_column_header(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Venta"]
        all_vals = [
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, 40) for c in range(1, 6)
        ]
        assert any("IGV" in v for v in all_vals)

    def test_costos_sheet_has_5_col_width(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Costos"]
        assert ws.column_dimensions["E"].width > 0

    def test_venta_sheet_has_5_col_width(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Venta"]
        assert ws.column_dimensions["E"].width > 0

    def test_costos_igv_value_for_visto_bueno(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Costos"]
        found = any(
            isinstance(ws.cell(row=r, column=4).value, (int, float))
            and abs(ws.cell(row=r, column=4).value - 8.10) < 0.01
            for r in range(1, 30)
        )
        assert found, "IGV value 8.10 for Visto Bueno (45 * 0.18) not in Costos sheet"

    def test_costos_subtotal_flete_header_present(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Costos"]
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 30)]
        assert any("Subtotal" in v for v in values)

    def test_venta_subtotal_flete_header_present(self):
        wb = _load_wb(_sintad_quote())
        ws = wb["Venta"]
        values = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 40)]
        assert any("Subtotal" in v for v in values)
