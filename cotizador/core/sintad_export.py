"""
SINTAD Excel export.

Abel re-enters every approved quote into SINTAD manually
(Meeting 3, 1:04:26). This module pre-fills every field
so the ejecutivo only has to copy-paste, not re-type.

4 sheets:
  1. Datos Generales  — shipment master data
  2. Costos           — internal cost breakdown (costeo)
  3. Venta            — sell-side line items
  4. Staff            — role assignments per SINTAD workflow

Staff logic (from Meeting 3, 1:16:09):
  Ejecutivo comercial : Jean Paul (GT-PC / GT-WCA) | Renato (others)
  Customer service    : Paulo Díaz (always)  # TODO confirm with Abel: Paolo / Pablo / Paulo — flagged 2026-06-09 demo
  Operativo           : Robin Lujan (imports) | Junior Loa (exports)
  Supervisor          : Kristel (always)
"""

from __future__ import annotations

import io
import json
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Brand colours ────────────────────────────────────────────────────────────
_ORANGE = "E8471C"
_NAVY   = "1B3A6B"
_GREY   = "D9D9D9"
_WHITE  = "FFFFFF"
_BLACK  = "000000"


def _hdr_fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def _bold(size: int = 11, colour: str = _BLACK) -> Font:
    return Font(bold=True, size=size, color=colour)


def _thin_border() -> Border:
    s = Side(style="thin", color=_BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_kv(ws, start_row: int, pairs: list[tuple[str, object]]) -> int:
    """Write (label, value) rows. Returns next available row."""
    border = _thin_border()
    for label, value in pairs:
        c_label = ws.cell(row=start_row, column=1, value=label)
        c_label.font   = _bold()
        c_label.fill   = _hdr_fill(_GREY)
        c_label.border = border
        c_label.alignment = Alignment(wrap_text=True)

        c_val = ws.cell(row=start_row, column=2, value=value)
        c_val.border    = border
        c_val.alignment = Alignment(wrap_text=True)

        start_row += 1
    return start_row


def _write_table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    """Write a header row + data rows. Returns next available row."""
    border = _thin_border()
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font   = Font(bold=True, color=_WHITE, size=10)
        c.fill   = _hdr_fill(_NAVY)
        c.border = border
        c.alignment = Alignment(horizontal="center")
    start_row += 1

    for row in rows:
        for col, val in enumerate(row, 1):
            c = ws.cell(row=start_row, column=col, value=val)
            c.border = border
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right")
        start_row += 1

    return start_row


def _write_section_header(ws, row: int, label: str, n_cols: int) -> int:
    """Write an orange merged section-header row. Returns next row."""
    border      = _thin_border()
    end_col_ltr = chr(ord("A") + n_cols - 1)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, color=_WHITE, size=10)
    c.fill      = _hdr_fill(_ORANGE)
    c.border    = border
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A{row}:{end_col_ltr}{row}")
    return row + 1


def _write_subtotal_row(ws, row: int, label: str, values: list, n_cols: int) -> int:
    """Write a grey subtotal row; values right-fill to the last len(values) cols."""
    border     = _thin_border()
    label_cols = n_cols - len(values)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, size=10)
    c.fill      = _hdr_fill(_GREY)
    c.border    = border
    if label_cols > 1:
        ws.merge_cells(f"A{row}:{chr(ord('A') + label_cols - 1)}{row}")
    for i, val in enumerate(values, label_cols + 1):
        cv = ws.cell(row=row, column=i, value=val)
        cv.font      = Font(bold=True, size=10)
        cv.fill      = _hdr_fill(_GREY)
        cv.border    = border
        cv.alignment = Alignment(horizontal="right")
    return row + 1


def _parse(raw) -> dict:
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw:
        try:
            return json.loads(raw)
        except Exception:
            pass
    return {}


def _is_export(quote: dict) -> bool:
    origin = (quote.get("origin") or "").lower()
    return any(k in origin for k in ["lima", "peru", "callao", "lim"])


def _staff(quote: dict) -> dict:
    staff_code = (quote.get("staff_code") or "GT-PC").upper()
    is_export  = _is_export(quote)

    ejecutivo = "Jean Paul" if staff_code in ("GT-PC", "GT-WCA") else "Renato"
    operativo = "Junior Loa" if is_export else "Robin Lujan"

    return {
        "Ejecutivo Comercial": ejecutivo,
        "Customer Service":    "Paulo Díaz",  # TODO confirm with Abel: Paolo / Pablo / Paulo — flagged 2026-06-09 demo
        "Operativo":           operativo,
        "Supervisor":          "Kristel",
    }


def generate_sintad_excel(quote: dict) -> bytes:
    """
    Build a SINTAD pre-fill Excel workbook from a quote dict.
    Returns raw bytes suitable for a Flask send_file() response.
    """
    costeo = _parse(quote.get("costeo_json"))
    venta  = _parse(quote.get("venta_json"))
    dims   = _parse(quote.get("dimensions_json"))

    ref        = quote.get("reference_code") or ""
    today      = date.today().strftime("%d/%m/%Y")
    validity   = venta.get("validity_days", 15)
    exc_rate   = quote.get("exchange_rate") or 0.0
    mode       = (quote.get("mode") or "lcl").upper()
    incoterm   = (quote.get("incoterm") or "").upper()
    is_export  = _is_export(quote)
    direction  = "Exportación" if is_export else "Importación"

    # Cargo type inference
    cargo_desc = (quote.get("cargo_description") or "").lower()
    if "peligrosa" in cargo_desc or "hazmat" in cargo_desc:
        cargo_type = "Mercancía Peligrosa"
    elif "perecible" in cargo_desc or "refrigerado" in cargo_desc:
        cargo_type = "Refrigerada / Temperatura controlada"
    elif mode == "FCL":
        cargo_type = "Contenedor (FCL)"
    elif mode == "LCL":
        cargo_type = "Carga Consolidada (LCL)"
    else:
        cargo_type = "Carga Aérea"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Datos Generales ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Datos Generales"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40

    # Title
    title = ws1.cell(row=1, column=1,
                     value=f"SINTAD — Datos Generales: {ref}")
    title.font = Font(bold=True, size=13, color=_WHITE)
    title.fill = _hdr_fill(_NAVY)
    ws1.merge_cells("A1:B1")
    title.alignment = Alignment(horizontal="center")

    sub = ws1.cell(row=2, column=1,
                   value=f"Generado: {today} · GT Cotizador v1.0")
    sub.font  = Font(italic=True, size=9, color=_NAVY)
    ws1.merge_cells("A2:B2")

    requester_label = (quote.get("requester_type") or "cliente").capitalize()
    _write_kv(ws1, 4, [
        ("Referencia",            ref),
        ("Cliente",               quote.get("client_name") or ""),
        ("Tipo Solicitante",      requester_label),   # row 6 — dropdown added below
        ("Tipo",                  direction),
        ("Puerto Origen",         quote.get("origin") or ""),
        ("Puerto Destino",        quote.get("destination") or ""),
        ("Incoterm",              incoterm),
        ("Modo de Transporte",    mode),
        ("Tipo de Carga",         cargo_type),
        ("Descripción Mercancía", quote.get("cargo_description") or ""),
        ("Peso (kg)",             quote.get("weight_kg") or 0.0),
        ("Volumen (CBM)",         round(quote.get("volume_cbm") or 0.0, 4)),
        ("Consolidador",          costeo.get("consolidator") or "—"),
        ("Agente de Aduana",      costeo.get("customs_agent") or "—"),
        ("Ruta",                  "Directa"),
        ("Días de Tránsito",      "TBD"),
        ("Frecuencia",            "Semanal"),
        ("Validez (días)",        validity),
        ("Tipo de Cambio SBS",    exc_rate),
        ("Fecha Cotización",      today),
    ])

    # Dropdown on Tipo Solicitante value cell (B6 = row 4 start + index 2)
    dv = DataValidation(type="list", formula1='"Agente,Cliente"', allow_blank=False, showErrorMessage=False)
    ws1.add_data_validation(dv)
    dv.add("B6")

    # ── Sheet 2: Costos ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Costos")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16

    t = ws2.cell(row=1, column=1, value=f"Costos — {ref}")
    t.font = Font(bold=True, size=13, color=_WHITE)
    t.fill = _hdr_fill(_ORANGE)
    ws2.merge_cells("A1:E1")
    t.alignment = Alignment(horizontal="center")

    note = ws2.cell(row=2, column=1,
                    value="⚠ SOLO INTERNO — No compartir con el cliente")
    note.font = Font(bold=True, color=_ORANGE)
    ws2.merge_cells("A2:E2")

    flete_val    = round(costeo.get("flete_internacional_usd") or 0, 2)
    flete_rate   = costeo.get("flete_rate_lcl") or 0
    flete_factor = costeo.get("flete_factor")
    thc_usd_val  = round(costeo.get("thc_usd") or 0, 2)
    thc_rate_val = costeo.get("thc_rate") or 0
    tn_m3_val    = round(flete_factor, 4) if flete_factor else ""

    # ── Costos Section 1: Flete Internacional ────────────────────────────
    r = _write_section_header(ws2, 4, "COSTOS DE FLETE INTERNACIONAL", 5)
    intl_costeo_rows = [
        ["Flete Internacional (USD)", flete_rate if flete_rate else flete_val, tn_m3_val, flete_val, ""],
    ]
    if thc_usd_val:
        intl_costeo_rows.append([
            "THC / Terminal Handling (USD)",
            thc_rate_val if thc_rate_val else thc_usd_val,
            tn_m3_val, thc_usd_val, "",
        ])
    for ei in (costeo.get("extra_items") or []):
        ei_tn = round(ei["factor"], 4) if ei.get("factor") else ""
        intl_costeo_rows.append([ei["concept"], round(ei["valor"], 2), ei_tn, round(ei["total"], 2), ""])
    r = _write_table(ws2, r, ["Concepto", "Costo", "TN/M3", "Total (USD)", ""], intl_costeo_rows)
    intl_subtotal = flete_val + thc_usd_val + sum(round(ei["total"], 2) for ei in (costeo.get("extra_items") or []))
    r = _write_subtotal_row(ws2, r, "Subtotal Flete Internacional", [round(intl_subtotal, 2), ""], 5)

    # ── Costos Section 2: Gastos Locales ─────────────────────────────────
    r += 1  # gap row
    r = _write_section_header(ws2, r, "GASTOS LOCALES (+ IGV 18%)", 5)
    vb_val      = round(costeo.get("visto_bueno_usd") or 0, 2)
    cust_val    = round(costeo.get("customs_agent_usd") or 0, 2)
    transp_val  = round(costeo.get("transport_usd") or 0, 2)
    aereo_val   = round(costeo.get("handling_aereo_usd") or 0, 2)
    local_costeo_rows = []
    if vb_val:
        local_costeo_rows.append(["Visto Bueno (USD)", vb_val, "", round(vb_val * 0.18, 2), round(vb_val * 1.18, 2)])
    if cust_val:
        local_costeo_rows.append(["Agente de Aduana (USD)", cust_val, "", round(cust_val * 0.18, 2), round(cust_val * 1.18, 2)])
    if aereo_val:
        local_costeo_rows.append(["Handling Aéreo (USD)", aereo_val, "", round(aereo_val * 0.18, 2), round(aereo_val * 1.18, 2)])
    if transp_val:
        local_costeo_rows.append(["Transporte Local (USD)", transp_val, "", round(transp_val * 0.18, 2), round(transp_val * 1.18, 2)])
    local_costeo_rows.append(["Transporte Local (S/)", round(costeo.get("transport_soles") or 0, 2), "", "", ""])
    r = _write_table(ws2, r, ["Concepto", "Monto Neto", "", "IGV 18%", "Total + IGV"], local_costeo_rows)
    local_neto = vb_val + cust_val + aereo_val + transp_val
    r = _write_subtotal_row(ws2, r, "Subtotal Gastos Locales",
                            [round(local_neto, 2), "", round(local_neto * 0.18, 2), round(local_neto * 1.18, 2)], 5)

    # ── Costos grand total + exchange rate ───────────────────────────────
    r += 1
    _write_kv(ws2, r, [
        ("TOTAL COSTEO (USD)", round(costeo.get("total_usd") or 0, 2)),
        ("Tipo de Cambio SBS", exc_rate),
    ])

    # ── Sheet 3: Venta ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Venta")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 16

    t3 = ws3.cell(row=1, column=1, value=f"Venta — {ref}")
    t3.font = Font(bold=True, size=13, color=_WHITE)
    t3.fill = _hdr_fill(_NAVY)
    ws3.merge_cells("A1:E1")
    t3.alignment = Alignment(horizontal="center")

    line_items = venta.get("line_items", [])
    intl_items = [i for i in line_items if not i.get("is_local")]
    local_items = [i for i in line_items if i.get("is_local")]
    has_factor  = any(i.get("factor_value") is not None for i in intl_items)

    # ── Venta Section 1: Flete Internacional ────────────────────────────
    rv = _write_section_header(ws3, 3, "COSTOS DE FLETE INTERNACIONAL", 5)
    if has_factor:
        intl_venta_rows = []
        for item in intl_items:
            if item.get("factor_value") is not None:
                fv = item.get("factor_value", 0)
                intl_venta_rows.append([
                    item.get("description", ""),
                    item.get("unit_rate", 0),
                    f"{fv:.4g}" if fv else "",
                    item.get("total", 0), "",
                ])
            else:
                intl_venta_rows.append([
                    item.get("description", ""), item.get("unit_price", 0), "",
                    item.get("total", 0), "",
                ])
        rv = _write_table(ws3, rv, ["Concepto", "Tarifa", "TN/M3", "Total (USD)", ""], intl_venta_rows)
    else:
        intl_venta_rows = [
            [i.get("description", ""), i.get("quantity", 1), i.get("unit_price", 0), i.get("total", 0), ""]
            for i in intl_items
        ]
        rv = _write_table(ws3, rv, ["Concepto", "Cant.", "Precio Unit. (USD)", "Total (USD)", ""], intl_venta_rows)
    intl_subtotal_v = sum(i.get("total") or 0 for i in intl_items)
    rv = _write_subtotal_row(ws3, rv, "Subtotal Flete Internacional", [round(intl_subtotal_v, 2), ""], 5)

    # ── Venta Section 2: Gastos Locales ──────────────────────────────────
    rv += 1
    rv = _write_section_header(ws3, rv, "GASTOS LOCALES (+ IGV 18%)", 5)
    local_venta_rows = []
    for item in local_items:
        neto = item.get("total") or 0
        local_venta_rows.append([
            item.get("description", ""),
            round(neto, 2),
            "",
            round(neto * 0.18, 2),
            round(neto * 1.18, 2),
        ])
    rv = _write_table(ws3, rv, ["Concepto", "Monto Neto", "", "IGV 18%", "Total + IGV"], local_venta_rows)
    local_neto_v = sum(i.get("total") or 0 for i in local_items)
    rv = _write_subtotal_row(ws3, rv, "Subtotal Gastos Locales",
                             [round(local_neto_v, 2), "", round(local_neto_v * 0.18, 2), round(local_neto_v * 1.18, 2)], 5)

    # ── Grand total + margin info ─────────────────────────────────────────
    rv += 1
    grand_total_v = intl_subtotal_v + local_neto_v * 1.18 if local_items else intl_subtotal_v + local_neto_v
    _write_subtotal_row(ws3, rv, "TOTAL VENTA (USD)", [round(grand_total_v, 2)], 5)
    next_r = rv + 2

    info = ws3.cell(row=next_r, column=1,
                    value=f"Margen: {(venta.get('margin_pct') or 0) * 100:.1f}%"
                          f"  ·  Validez: {validity} días")
    info.font = Font(italic=True, size=9, color=_NAVY)
    ws3.merge_cells(f"A{next_r}:E{next_r}")

    # ── Sheet 4: Staff ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Staff")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 25

    t4 = ws4.cell(row=1, column=1, value=f"Asignación de Staff — {ref}")
    t4.font = Font(bold=True, size=13, color=_WHITE)
    t4.fill = _hdr_fill(_NAVY)
    ws4.merge_cells("A1:B1")
    t4.alignment = Alignment(horizontal="center")

    staff = _staff(quote)
    staff_rows = [[role, name] for role, name in staff.items()]
    next_r4 = _write_table(ws4, 3, ["Rol", "Nombre"], staff_rows)

    dir_note = ws4.cell(row=next_r4 + 1, column=1,
                        value=f"Dirección detectada: {direction}")
    dir_note.font = Font(italic=True, size=9, color=_NAVY)
    ws4.merge_cells(f"A{next_r4+1}:B{next_r4+1}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
